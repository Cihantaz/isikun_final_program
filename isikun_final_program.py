# power2.py
# Flask Final Exam Scheduler - HTML-compatible build
# - Type-safe parsing everywhere (prevents str/int compare errors)
# - Greedy preview + CP-SAT solve (OR-Tools) with UNASSIGNED (always feasible)
# - Excel/CSV import hardened (regex day parsing fixes 1.6 -> [1,6])
# - Backward compatible endpoints (bulk_set_caps, save_ics, etc.)
#
# Requirements:
#   pip install flask openpyxl pandas ortools
#
# Run:
#   python power2.py
# Open:
#   http://127.0.0.1:5000

from __future__ import annotations

from dataclasses import dataclass, field

import csv
import io
import math
import os
import re
from collections import defaultdict
from datetime import datetime, timedelta, time
from types import SimpleNamespace
from typing import Any, Dict, List, Optional, Tuple

from flask import Flask, flash, redirect, render_template, request, send_file, url_for

try:
    import openpyxl  # type: ignore
except Exception as e:
    raise RuntimeError("openpyxl gerekli: pip install openpyxl") from e

try:
    import pandas as pd  # type: ignore
    HAS_PANDAS = True
except Exception:
    HAS_PANDAS = False

try:
    from ortools.sat.python import cp_model  # type: ignore
    from ortools.sat.python.cp_model import LinearExpr  # type: ignore
    HAS_ORTOOLS = True
except Exception:
    HAS_ORTOOLS = False


# =========================
# Constants
# =========================

MAX_CAP = 10 ** 9  # Kapasite limiti için varsayılan "sonsuzluk" değeri

# =========================
# Type-safe helpers
# =========================

DAY_RE = re.compile(r"\d+")


def to_int(v: Any, default: int = 0) -> int:
    try:
        if v is None:
            return default
        if isinstance(v, bool):
            return int(v)
        if isinstance(v, (int, float)):
            return int(v)
        s = str(v).strip()
        if s == "":
            return default
        return int(float(s))
    except Exception:
        return default


def to_opt_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None
    try:
        return to_int(s, default=None)  # type: ignore[arg-type]
    except Exception:
        return None


def clamp_int(v: Any, lo: int, hi: int) -> Optional[int]:
    x = to_opt_int(v)
    if x is None:
        return None
    if x < lo:
        return lo
    if x > hi:
        return hi
    return x


def to_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    if isinstance(v, int):
        return v != 0
    s = str(v).strip().lower()
    return s in ("1", "true", "yes", "on", "y", "t", "x")


def norm_code(v: Any) -> str:
    return (str(v).strip().upper() if v is not None else "").strip()


def base_code(code: str) -> str:
    return norm_code(code).replace("__EXTRA__", "")


def parse_days(value: Any, n_days: int) -> List[int]:
    if value is None:
        return []
    nums = [to_int(x, 0) for x in DAY_RE.findall(str(value))]
    out = sorted({d for d in nums if 1 <= d <= n_days})
    return out


def parse_ids(value: Any) -> List[str]:
    if value is None:
        return []
    s = str(value).strip()
    if not s:
        return []
    parts = re.findall(r"[A-Za-z0-9_]+", s)
    return [p.strip() for p in parts if p.strip()]


def hhmm_to_time(s: str) -> Optional[time]:
    s = (s or "").strip()
    m = re.match(r"^(\d{1,2}):(\d{2})$", s)
    if not m:
        return None
    hh = to_int(m.group(1), 0)
    mm = to_int(m.group(2), 0)
    if not (0 <= hh <= 23 and 0 <= mm <= 59):
        return None
    return time(hh, mm)


def time_to_hhmm(t: time) -> str:
    return f"{t.hour:02d}:{t.minute:02d}"


def day_slot_times(slot: int, day_start: time, slot_len: int) -> Tuple[time, time]:
    start_min = (slot - 1) * slot_len
    sh = day_start.hour * 60 + day_start.minute + start_min
    eh = sh + slot_len
    st = time(sh // 60, sh % 60)
    en = time(eh // 60, eh % 60)
    return st, en


def period_index(day: int, slot: int, spd: int) -> int:
    return (day - 1) * spd + slot


def period_to_day_slot(p: int, spd: int) -> Tuple[int, int]:
    d = (p - 1) // spd + 1
    s = (p - 1) % spd + 1
    return d, s


def make_heat(load: int, denom: int) -> str:
    denom = max(1, to_int(denom, 1))
    r = max(0.0, min(1.0, load / denom))
    red = 255
    green = int(255 * (1.0 - r))
    blue = int(255 * (1.0 - r))
    return f"rgb({red},{green},{blue})"


# =========================
# App / State
# =========================

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET", "dev-secret-change-me")


@dataclass
class AppState:
    calendar: Dict[str, Any] = field(default_factory=lambda: {
        "n_days": 10,
        "slots_per_day": 3,
        "slot_length_min": 180,
        "buffer_minutes": 0,
        "day_start_time": "08:30",
        "day_end_time": "18:00",
        "start_date": "2026-06-08",
        "end_date": "2026-06-21",
        "last_slot_forbidden_for_three_hour": False,
    })

    courses: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    conflicts: List[Tuple[str, str, int]] = field(default_factory=list)
    group_constraints: List[Dict[str, Any]] = field(default_factory=list)
    instructors: Dict[str, Dict[str, str]] = field(default_factory=dict)
    instructor_unavailability: List[Dict[str, Any]] = field(default_factory=list)

    slot_caps: Dict[Tuple[int, int], Dict[str, int]] = field(
        default_factory=lambda: defaultdict(lambda: {"min": 0, "max": MAX_CAP})
    )
    seeds: List[Dict[str, Any]] = field(default_factory=list)

    preview: Dict[str, Tuple[int, int]] = field(default_factory=dict)
    preview_unassigned: List[Tuple[str, str, List[Tuple[int, int]]]] = field(default_factory=list)
    final: Dict[str, Tuple[int, int]] = field(default_factory=dict)
    final_unassigned: List[Tuple[str, str, List[Tuple[int, int]]]] = field(default_factory=list)
    last_diagnostics: List[str] = field(default_factory=list)
    kontrol_results: Dict[str, Any] = field(default_factory=dict)
    degistir_onerileri: List[Dict[str, Any]] = field(default_factory=list)


appstate = AppState()


def reset_results() -> None:
    appstate.preview.clear()
    appstate.preview_unassigned.clear()
    appstate.final.clear()
    appstate.final_unassigned.clear()


def reset_all_state() -> None:
    global appstate
    appstate = AppState()
    reset_results()


def same_tab(tab: str) -> Any:
    return redirect(url_for("index", tab=tab))


# =========================
# Template wrappers (HTML expects dot-attributes)
# =========================

def cal_obj() -> SimpleNamespace:
    c = appstate.calendar
    return SimpleNamespace(
        n_days=to_int(c.get("n_days"), 9),
        slots_per_day=to_int(c.get("slots_per_day"), 4),
        slot_length_min=to_int(c.get("slot_length_min"), 60),
        buffer_minutes=to_int(c.get("buffer_minutes"), 0),
        day_start_time=str(c.get("day_start_time") or "08:30"),
        day_end_time=str(c.get("day_end_time") or "18:00"),
        start_date=c.get("start_date"),
        end_date=c.get("end_date"),
        last_slot_forbidden_for_three_hour=to_bool(c.get("last_slot_forbidden_for_three_hour")),
    )


def course_obj(code: str, d: Dict[str, Any]) -> SimpleNamespace:
    return SimpleNamespace(
        code=code,
        name=str(d.get("name") or ""),
        duration=to_int(d.get("duration"), to_int(appstate.calendar.get("slot_length_min"), 60)),
        size=to_int(d.get("size"), 0),
        preferred_days=list(d.get("preferred_days") or []),
        forbidden_days=list(d.get("forbidden_days") or []),
        fixed_day=d.get("fixed_day"),
        fixed_slot=d.get("fixed_slot"),
        multi_slots=to_int(d.get("multi_slots"), 1),
        three_hour=to_bool(d.get("three_hour")),
        instructors=list(d.get("instructors") or []),
    )


def instructors_obj() -> Dict[str, SimpleNamespace]:
    out: Dict[str, SimpleNamespace] = {}
    for iid, info in appstate.instructors.items():
        out[iid] = SimpleNamespace(name=str(info.get("name") or ""), email=str(info.get("email") or ""))
    return out


def unavailability_obj() -> List[SimpleNamespace]:
    return [
        SimpleNamespace(instr=str(u.get("instr") or ""), day=to_int(u.get("day"), 0), slot=to_int(u.get("slot"), 0))
        for u in appstate.instructor_unavailability
    ]


def group_constraints_obj() -> List[SimpleNamespace]:
    return [
        SimpleNamespace(type=str(g.get("type") or ""), courses=list(g.get("courses") or []))
        for g in appstate.group_constraints
    ]


def slot_caps_obj() -> Dict[Tuple[int, int], SimpleNamespace]:
    out: Dict[Tuple[int, int], SimpleNamespace] = {}
    n = to_int(appstate.calendar.get("n_days"), 9)
    spd = to_int(appstate.calendar.get("slots_per_day"), 4)
    for d in range(1, n + 1):
        for s in range(1, spd + 1):
            cap = appstate.slot_caps.get((d, s), {"min": 0, "max": MAX_CAP})
            mn = to_int(cap.get("min"), 0)
            mx = to_int(cap.get("max"), MAX_CAP)
            if mn > mx:
                mn, mx = mx, mn
            out[(d, s)] = SimpleNamespace(min=mn, max=mx)
    return out


def seeds_obj() -> List[SimpleNamespace]:
    return [
        SimpleNamespace(course=norm_code(s.get("course")), day=to_int(s.get("day"), 0),
                       slot=to_int(s.get("slot"), 0), lock=to_bool(s.get("lock")))
        for s in appstate.seeds
    ]


# =========================
# Importers
# =========================

COURSE_COLS = {
    "ders kodu": "code",
    "ders adı": "name",
    "süre (dk)": "duration",
    "öğrenci sayısı": "size",
    "tercih günleri": "preferred_days",
    "engelli günler": "forbidden_days",
    "sabit gün": "fixed_day",
    "sabit slot": "fixed_slot",
    "multi slots": "multi_slots",
    "3 saatlik": "three_hour",
    "eğitmen id’leri": "instructor_ids",
    "eğitmen id'leri": "instructor_ids",
    "eğitmen idleri": "instructor_ids",
    "egitmen idleri": "instructor_ids",
}


def load_courses_excel(data: bytes) -> None:
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    headers = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[1]]
    colmap: Dict[int, str] = {}
    for idx, h in enumerate(headers):
        key = COURSE_COLS.get(h)
        if key:
            colmap[idx] = key

    n_days = to_int(appstate.calendar.get("n_days"), 9)
    spd = to_int(appstate.calendar.get("slots_per_day"), 4)
    slot_len = to_int(appstate.calendar.get("slot_length_min"), 60)

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        rowd: Dict[str, Any] = {}
        for cidx, key in colmap.items():
            if cidx < len(row):
                rowd[key] = row[cidx]

        code = norm_code(rowd.get("code"))
        if not code:
            continue

        name = str(rowd.get("name") or "")
        size = to_int(rowd.get("size"), 0)
        duration = to_int(rowd.get("duration"), slot_len)
        multi_slots = to_int(rowd.get("multi_slots"), max(1, int(math.ceil(duration / max(1, slot_len)))))

        three_hour = to_bool(rowd.get("three_hour"))
        if three_hour:
            duration = 180
            if "multi_slots" not in rowd or rowd.get("multi_slots") in (None, ""):
                multi_slots = max(1, int(math.ceil(duration / max(1, slot_len))))

        preferred = parse_days(rowd.get("preferred_days"), n_days)
        forbidden = parse_days(rowd.get("forbidden_days"), n_days)

        fixed_day = clamp_int(rowd.get("fixed_day"), 1, n_days)
        fixed_slot = clamp_int(rowd.get("fixed_slot"), 1, spd)

        instructors = parse_ids(rowd.get("instructor_ids"))

        appstate.courses[code] = {
            "code": code,
            "name": name,
            "duration": duration,
            "size": size,
            "preferred_days": preferred,
            "forbidden_days": forbidden,
            "fixed_day": fixed_day,
            "fixed_slot": fixed_slot,
            "multi_slots": multi_slots,
            "three_hour": three_hour,
            "instructors": instructors,
        }
        count += 1

    reset_results()
    flash(f"Excel yüklendi: {count} ders", "success")


def load_conflicts_csv(data: bytes) -> None:
    text = data.decode("utf-8", errors="ignore")
    f = io.StringIO(text)
    reader = csv.reader(f)
    rows = [r for r in reader if r and any(x.strip() for x in r)]
    if not rows:
        flash("CSV boş.", "warning")
        return

    header = [x.strip().lower() for x in rows[0]]
    start_idx = 0
    if any("course" in h or "ders" in h for h in header) or any("weight" in h for h in header):
        start_idx = 1

    added = 0
    for r in rows[start_idx:]:
        if len(r) < 2:
            continue
        a = norm_code(r[0])
        b = norm_code(r[1])
        w = to_int(r[2], 1) if len(r) >= 3 else 1
        if not a or not b or a == b:
            continue
        # Ignore conflicts that reference courses not present in the system
        if a not in appstate.courses or b not in appstate.courses:
            continue
        appstate.conflicts.append((a, b, max(1, w)))
        added += 1

    reset_results()
    flash(f"Çakışmalar yüklendi: {added}", "success")


def load_group_constraints_excel(data: bytes) -> None:
    if not HAS_PANDAS:
        raise RuntimeError("pandas gerekli: pip install pandas")
    df = pd.read_excel(io.BytesIO(data))  # type: ignore[name-defined]
    cols = list(df.columns)
    use_cols = cols[:4]
    first_col_name = str(cols[0]).strip().lower()
    if len(cols) >= 5 and (first_col_name.startswith('unnamed') or str(cols[0]).strip().isdigit()):
        use_cols = cols[1:5]

    mapping = {0: "SameDay", 1: "SameSlot", 2: "DifferentDay", 3: "DifferentSlot"}

    added = 0
    for _, row in df.iterrows():
        for i, col in enumerate(use_cols):
            v = row.get(col)
            if v is None or str(v).strip() == "":
                continue
            codes = [norm_code(x) for x in re.split(r"[,\s;]+", str(v)) if norm_code(x)]
            codes = [c for c in codes if c in appstate.courses]
            if len(codes) >= 2:
                appstate.group_constraints.append({"type": mapping[i], "courses": codes})
                added += 1

    reset_results()
    flash(f"Grup kısıtları yüklendi: {added}", "success")


# =========================
# Greedy preview
# =========================

def greedy_preview() -> None:
    reset_results()
    n_days = to_int(appstate.calendar.get("n_days"), 9)
    spd = to_int(appstate.calendar.get("slots_per_day"), 4)

    conf_pairs = set()
    for a, b, _w in appstate.conflicts:
        a2, b2 = norm_code(a), norm_code(b)
        if a2 and b2:
            conf_pairs.add((a2, b2))
            conf_pairs.add((b2, a2))

    for g in appstate.group_constraints:
        if str(g.get("type") or "") == "DifferentSlot":
            L = [norm_code(x) for x in (g.get("courses") or []) if norm_code(x)]
            L = [x for x in L if x in appstate.courses]
            for i in range(len(L)):
                for j in range(i + 1, len(L)):
                    conf_pairs.add((L[i], L[j]))
                    conf_pairs.add((L[j], L[i]))

    slot_load: Dict[Tuple[int, int], int] = defaultdict(int)

    def fits(code: str, d: int, s: int) -> Tuple[bool, str]:
        cd = appstate.courses.get(code, {})
        size = to_int(cd.get("size"), 0)

        forb = [to_int(x, 0) for x in (cd.get("forbidden_days") or [])]
        if d in forb:
            return False, "Engelli gün"

        fd = to_int(cd.get("fixed_day"), 0)
        fs = to_int(cd.get("fixed_slot"), 0)
        if fd > 0 and fs > 0 and (d, s) != (fd, fs):
            return False, "Sabit"

        instrs = [str(x).strip() for x in (cd.get("instructors") or []) if str(x).strip()]
        if instrs:
            for un in appstate.instructor_unavailability:
                if str(un.get("instr") or "") in instrs and to_int(un.get("day"), 0) == d and to_int(un.get("slot"), 0) == s:
                    return False, "Eğitmen"

        cap = appstate.slot_caps.get((d, s), {"min": 0, "max": MAX_CAP})
        capmax = to_int(cap.get("max"), MAX_CAP)
        if capmax < MAX_CAP and slot_load[(d, s)] + size > capmax:
            return False, "Kapasite"

        for other, (od, os) in appstate.preview.items():
            other_base = base_code(other)
            if (od, os) == (d, s) and (code, other_base) in conf_pairs:
                return False, "Çakışma/Grup"
        return True, ""

    order = sorted(
        appstate.courses.keys(),
        key=lambda c: (
            0 if (to_int(appstate.courses[c].get("fixed_day"), 0) > 0 and to_int(appstate.courses[c].get("fixed_slot"), 0) > 0) else 1,
            -to_int(appstate.courses[c].get("size"), 0),
            c,
        ),
    )

    for code in order:
        cd = appstate.courses[code]
        size = to_int(cd.get("size"), 0)
        pref = [to_int(x, 0) for x in (cd.get("preferred_days") or [])]
        pref = [d for d in pref if 1 <= d <= n_days]

        candidates: List[Tuple[int, int]] = []
        reason = "Uygun slot yok"
        days = pref + [d for d in range(1, n_days + 1) if d not in pref]
        for d in days:
            for s in range(1, spd + 1):
                ok, why = fits(code, d, s)
                if ok:
                    candidates.append((d, s))
                elif why:
                    reason = why

        placed = False
        for (d, s) in candidates:
            appstate.preview[code] = (d, s)
            slot_load[(d, s)] += size
            placed = True
            break

        if not placed:
            alts = candidates[:3]
            appstate.preview_unassigned.append((code, reason, alts))


# =========================
# Helper: cached conflict / group pairs
# =========================

def get_conflict_pairs() -> set[Tuple[str, str]]:
    """Sistemdeki çakışma çiftlerini normalize edilmiş şekilde döndürür."""
    pairs: set[Tuple[str, str]] = set()
    for a, b, _w in appstate.conflicts:
        a2, b2 = norm_code(a), norm_code(b)
        if a2 and b2 and a2 != b2:
            pairs.add((a2, b2))
            pairs.add((b2, a2))
    return pairs


def get_group_pairs() -> Tuple[
    set[Tuple[str, str]], set[Tuple[str, str]], set[Tuple[str, str]], set[Tuple[str, str]]
]:
    """Grup kısıtlarından (SameDay, DiffDay, SameSlot, DiffSlot) çiftleri döndürür."""
    same_day: set[Tuple[str, str]] = set()
    diff_day: set[Tuple[str, str]] = set()
    same_slot: set[Tuple[str, str]] = set()
    diff_slot: set[Tuple[str, str]] = set()
    for g in appstate.group_constraints:
        gtype = str(g.get("type") or "")
        members = [norm_code(x) for x in (g.get("courses") or [])]
        members = [m for m in members if m in appstate.courses]
        if len(members) < 2:
            continue
        for i in range(len(members)):
            for j in range(i + 1, len(members)):
                a, b = members[i], members[j]
                key = (a, b) if a < b else (b, a)
                if gtype == "SameDay":
                    same_day.add(key)
                elif gtype == "DifferentDay":
                    diff_day.add(key)
                elif gtype == "SameSlot":
                    same_slot.add(key)
                elif gtype == "DifferentSlot":
                    diff_slot.add(key)
    return same_day, diff_day, same_slot, diff_slot


# =========================
# Feasibility Doctor — alt kontrol fonksiyonları
# =========================

def _check_calendar_issues(notes: List[str], cal: Dict[str, Any]) -> None:
    spd = to_int(cal.get("slots_per_day"), 4)
    slot_len = to_int(cal.get("slot_length_min"), 60)
    st = hhmm_to_time(str(cal.get("day_start_time") or "08:30")) or time(8, 30)
    en = hhmm_to_time(str(cal.get("day_end_time") or "18:00")) or time(18, 0)
    day_total_min = (en.hour * 60 + en.minute) - (st.hour * 60 + st.minute)
    slots_total_min = spd * slot_len
    if slots_total_min > day_total_min:
        notes.append(
            f"TAKVİM UYARISI: Gün başlangıç-bitiş aralığı {day_total_min} dk, "
            f"ama {spd} slot × {slot_len} dk = {slots_total_min} dk gerekiyor. Slotlar güne sığmıyor."
        )


def _check_course_issues(
    notes: List[str], courses: Dict[str, Dict[str, Any]],
    n_days: int, spd: int, last_forbidden: bool, all_instr_ids: set[str]
) -> int:
    total_students = 0
    for code, cd in courses.items():
        ms = to_int(cd.get("multi_slots"), 1)
        fd = to_int(cd.get("fixed_day"), 0)
        fs = to_int(cd.get("fixed_slot"), 0)
        pref = [to_int(x, 0) for x in (cd.get("preferred_days") or [])]
        forb = [to_int(x, 0) for x in (cd.get("forbidden_days") or [])]
        three = to_bool(cd.get("three_hour"))
        size = to_int(cd.get("size"), 0)
        total_students += size
        instrs = [str(i).strip() for i in (cd.get("instructors") or []) if str(i).strip()]

        if ms > spd:
            notes.append(f"{code}: multi_slots ({ms}) > gün başına slot ({spd}). Bu ders hiçbir güne sığmaz.")
        if fs > 0 and fs + ms - 1 > spd:
            notes.append(f"{code}: Sabit slot S{fs} + multi_slots ({ms}) gün sonunu aşıyor (S{spd}).")
        if fd > 0 and fd in forb:
            notes.append(f"{code}: Sabit gün G{fd} aynı zamanda engelli günler listesinde.")
        overlap = set(pref) & set(forb)
        if overlap:
            notes.append(f"{code}: Gün {sorted(overlap)} hem tercih edilen hem engelli olarak tanımlanmış.")
        for iid in instrs:
            if iid not in all_instr_ids:
                notes.append(f"{code}: Eğitmen ID '{iid}' kayıtlı eğitmenlerde bulunamadı.")
        if three and last_forbidden and fs == spd and ms >= 1:
            notes.append(
                f"{code}: 3 saatlik sınav son slot (S{spd}) için sabitlenmiş, "
                f"ancak takvim ayarında son slot 3 saatlik sınavlar için yasak."
            )
    return total_students


def _check_instructor_issues(notes: List[str], instructors: Dict[str, Any], unavailability: List[Dict[str, Any]]) -> None:
    for rec in unavailability:
        iid = str(rec.get("instr") or "").strip()
        if iid and iid not in instructors:
            d = to_int(rec.get("day"), 0)
            s = to_int(rec.get("slot"), 0)
            notes.append(f"Eğitmen uygun olmama kaydı (I={iid}, G{d}, S{s}): Bu ID kayıtlı eğitmenlerde yok.")


def _check_seed_issues(notes: List[str], courses: Dict[str, Any], seeds: List[Dict[str, Any]], n_days: int, spd: int) -> None:
    for s in seeds:
        c = norm_code(s.get("course"))
        sd = to_int(s.get("day"), 0)
        ssl = to_int(s.get("slot"), 0)
        lock = to_bool(s.get("lock"))
        if c not in courses:
            notes.append(f"Seed ({c} -> G{sd}/S{ssl}): Bu ders kodu kayıtlı derslerde yok.")
            continue
        cd = courses[c]
        fd = to_int(cd.get("fixed_day"), 0)
        fs = to_int(cd.get("fixed_slot"), 0)
        forb = [to_int(x, 0) for x in (cd.get("forbidden_days") or [])]
        if lock and fd > 0 and fs > 0 and (sd, ssl) != (fd, fs):
            notes.append(
                f"Seed çelişkisi: {c} dersi sabit G{fd}/S{fs} olarak tanımlı, "
                f"ancak lock'lu seed G{sd}/S{ssl} konumunda. Seed sabit atamayı geçersiz kılar."
            )
        if sd in forb:
            notes.append(f"Seed uyarısı: {c} -> G{sd}/S{ssl} engelli günlerden birine denk geliyor.")
        if sd > n_days or ssl > spd:
            notes.append(f"Seed uyarısı: {c} -> G{sd}/S{ssl} takvim sınırlarını aşıyor.")


def _check_capacity_issues(
    notes: List[str], slot_caps: Dict[Tuple[int, int], Dict[str, int]],
    n_days: int, spd: int, courses: Dict[str, Any], total_students: int
) -> None:
    total_capmax = 0
    total_capmin = 0
    max_single_cap = 0
    for d in range(1, n_days + 1):
        for s in range(1, spd + 1):
            cap = slot_caps.get((d, s), {"min": 0, "max": MAX_CAP})
            mn = to_int(cap.get("min"), 0)
            mx = to_int(cap.get("max"), MAX_CAP)
            if mn > mx:
                notes.append(
                    f"Slot G{d}/S{s}: Min kapasite ({mn}) max kapasiteden ({mx}) büyük. "
                    f"Sistem otomatik yer değiştirdi, lütfen düzeltin."
                )
            total_capmax += mx if mx < MAX_CAP else 0
            total_capmin += mn
            if mx < MAX_CAP and mx > max_single_cap:
                max_single_cap = mx

    if total_capmax > 0 and total_students > total_capmax:
        notes.append(
            f"KAPASİTE UYARISI: Toplam öğrenci sayısı ({total_students}) > "
            f"toplam slot kapasitesi ({total_capmax}). Tüm dersler fiziksel olarak sığmaz."
        )
    if total_capmin > total_students:
        notes.append(
            f"KAPASİTE UYARISI: Toplam minimum kapasite ({total_capmin}) > "
            f"toplam öğrenci sayısı ({total_students}). Min kapasiteler karşılanamaz."
        )
    if max_single_cap > 0:
        for code, cd in courses.items():
            sz = to_int(cd.get("size"), 0)
            if sz > max_single_cap:
                notes.append(
                    f"{code}: Öğrenci sayısı ({sz}) tek slotun max kapasitesinden ({max_single_cap}) büyük. "
                    f"Bu ders tek başına hiçbir slota sığmaz."
                )


def _check_conflict_data_issues(notes: List[str], conflicts: List[Tuple[str, str, int]], course_codes: set[str]) -> None:
    for a, b, w in conflicts:
        a2, b2 = norm_code(a), norm_code(b)
        if a2 not in course_codes or b2 not in course_codes:
            continue
        if a2 == b2:
            notes.append(f"Çakışma kaydı ({a},{b}): Bir ders kendisiyle çakışamaz.")


def _check_group_data_issues(notes: List[str], group_constraints: List[Dict[str, Any]], course_codes: set[str]) -> None:
    for g in group_constraints:
        members = [norm_code(x) for x in (g.get("courses") or [])]
        missing = [m for m in members if m and m not in course_codes]
        if missing:
            notes.append(f"Grup kısıtı ({g.get('type')}): {', '.join(missing)} ders kodları sistemde yok.")
        seen = set()
        dups = set()
        for m in members:
            if m in seen:
                dups.add(m)
            seen.add(m)
        if dups:
            notes.append(f"Grup kısıtı ({g.get('type')}): {', '.join(dups)} tekrar eden ders kodu var.")


def _check_group_conflicts(
    notes: List[str], group_constraints: List[Dict[str, Any]],
    courses: Dict[str, Any], spd: int
) -> None:
    conflict_pairs = get_conflict_pairs()
    same_day, diff_day, same_slot, diff_slot = get_group_pairs()

    for a, b in same_slot:
        if (a, b) in conflict_pairs or (b, a) in conflict_pairs:
            notes.append(
                f"UYARI: {a} ve {b} dersleri çakışma listesinde olduğu halde SameSlot grubuna da eklenmiş. "
                f"Çakışma zaten aynı slotta olmalarını yasaklar. SameSlot grubu gereksizdir."
            )

    for key in same_day & diff_day:
        a, b = key
        notes.append(
            f"ÇELİŞKİ: {a} ve {b} dersleri için hem SameDay hem DifferentDay kısıtı tanımlanmış. "
            f"Aynı günde olmaları ve farklı günde olmaları aynı anda mümkün değil."
        )

    for key in same_slot & diff_slot:
        a, b = key
        notes.append(
            f"ÇELİŞKİ: {a} ve {b} dersleri için hem SameSlot hem DifferentSlot kısıtı tanımlanmış. "
            f"Aynı slotta olmaları ve farklı slotta olmaları aynı anda mümkün değil."
        )

    for g in group_constraints:
        gtype = str(g.get("type") or "")
        members = [norm_code(x) for x in (g.get("courses") or [])]
        members = [m for m in members if m in courses]
        if len(members) < 2:
            continue

        fixed_map: Dict[str, Tuple[int, int]] = {}
        for c in members:
            cd = courses[c]
            fd = to_int(cd.get("fixed_day"), 0)
            fs = to_int(cd.get("fixed_slot"), 0)
            if fd > 0 and fs > 0:
                fixed_map[c] = (fd, fs)

        if gtype == "SameSlot":
            periods = {c: period_index(d, s, spd) for c, (d, s) in fixed_map.items()}
            if len(set(periods.values())) > 1:
                details = ", ".join(f"{c}=G{d}S{s}" for c, (d, s) in fixed_map.items())
                notes.append(
                    f"SameSlot grubu çelişkisi: {details} — aynı slotta olmaları gerekirken farklı sabit slotlar verilmiş."
                )
        elif gtype == "SameDay":
            days = {c: d for c, (d, _s) in fixed_map.items()}
            if len(set(days.values())) > 1:
                details = ", ".join(f"{c}=G{d}" for c, d in days.items())
                notes.append(
                    f"SameDay grubu çelişkisi: {details} — aynı günde olmaları gerekirken farklı sabit günler verilmiş."
                )
        elif gtype == "DifferentSlot":
            seen_periods: Dict[int, List[str]] = defaultdict(list)
            for c, (d, s) in fixed_map.items():
                seen_periods[period_index(d, s, spd)].append(c)
            for per, codes in seen_periods.items():
                if len(codes) > 1:
                    d, s = period_to_day_slot(per, spd)
                    notes.append(
                        f"DifferentSlot grubu çelişkisi: {', '.join(codes)} hepsi G{d}S{s} — "
                        f"farklı slotta olmaları gerekirken aynı sabit slota yerleştirilmiş."
                    )
        elif gtype == "DifferentDay":
            seen_days: Dict[int, List[str]] = defaultdict(list)
            for c, (d, _s) in fixed_map.items():
                seen_days[d].append(c)
            for d, codes in seen_days.items():
                if len(codes) > 1:
                    notes.append(
                        f"DifferentDay grubu çelişkisi: {', '.join(codes)} hepsi G{d} — "
                        f"farklı günde olmaları gerekirken aynı sabit güne yerleştirilmiş."
                    )


def diagnose_all_issues() -> List[str]:
    """Tüm sistem verilerini tarar ve kullanıcıya gösterilecek uyarı/çelişki listesi döner.
    15+ farklı senaryo tespit edilir."""
    notes: List[str] = []
    cal = appstate.calendar
    n_days = to_int(cal.get("n_days"), 9)
    spd = to_int(cal.get("slots_per_day"), 4)
    slot_len = to_int(cal.get("slot_length_min"), 60)
    last_forbidden = to_bool(cal.get("last_slot_forbidden_for_three_hour"))

    _check_calendar_issues(notes, cal)
    total_students = _check_course_issues(notes, appstate.courses, n_days, spd, last_forbidden, set(appstate.instructors.keys()))
    _check_instructor_issues(notes, appstate.instructors, appstate.instructor_unavailability)
    _check_seed_issues(notes, appstate.courses, appstate.seeds, n_days, spd)
    _check_capacity_issues(notes, appstate.slot_caps, n_days, spd, appstate.courses, total_students)

    course_codes = set(appstate.courses.keys())
    _check_conflict_data_issues(notes, appstate.conflicts, course_codes)
    _check_group_data_issues(notes, appstate.group_constraints, course_codes)
    _check_group_conflicts(notes, appstate.group_constraints, appstate.courses, spd)

    return notes


def analyze_assignment_issues(assign: Dict[str, Tuple[int, int]], slot_load: Dict[Tuple[int, int], int]) -> List[str]:
    issues: List[str] = []
    n_days = to_int(appstate.calendar.get("n_days"), 9)
    spd = to_int(appstate.calendar.get("slots_per_day"), 4)

    placed = {base_code(k): (to_int(v[0], 0), to_int(v[1], 0)) for k, v in assign.items() if "__extra__" not in k}

    # Individual course assignment checks
    for code, (d, s) in placed.items():
        cd = appstate.courses.get(code, {})
        if not cd:
            continue
        fd = to_int(cd.get("fixed_day"), 0)
        fs = to_int(cd.get("fixed_slot"), 0)
        forb = [to_int(x, 0) for x in (cd.get("forbidden_days") or [])]
        pref = [to_int(x, 0) for x in (cd.get("preferred_days") or [])]
        instrs = [str(i).strip() for i in (cd.get("instructors") or []) if str(i).strip()]

        if fd > 0 and fs > 0 and (d, s) != (fd, fs):
            issues.append(
                f"{code}: Sabit gün/slot G{fd}/S{fs} olarak girildi, ancak G{d}/S{s} yerleşti. "
                f"Bu, sabit atama kısıtının sağlanamadığı veya başka kısıtların G{fd}/S{fs} konumunu engellediği anlamına gelir."
            )
        elif fd > 0 and d != fd:
            issues.append(
                f"{code}: Sabit gün G{fd} olarak girildi, ancak G{d} yerleşti. "
                f"Bu durumda G{fd} içinde uygun slot kalmamış veya diğer kısıtlar G{fd}'yi kullanılamaz hale getirmiş olabilir."
            )
        elif fs > 0 and s != fs:
            issues.append(
                f"{code}: Sabit slot S{fs} olarak girildi, ancak S{s} yerleşti. "
                f"Bu, S{fs} slotunda dersin yerleşmesini engelleyen başka bir kısıt olduğu anlamına gelir."
            )

        if d in forb:
            issues.append(
                f"{code}: Engelli gün G{d} üzerine yerleşti. "
                f"Bu dersin G{d} gününü alması yasaklanmış."
            )

        if pref and d not in pref:
            issues.append(
                f"{code}: Tercih edilen günler {pref} dışında G{d} yerleşti. "
                "Bu, tercih günlerinde uygun slot bulunmadığını gösterir."
            )

        for iid in instrs:
            unavailable = [to_int(rec.get("slot"), 0) for rec in appstate.instructor_unavailability
                           if str(rec.get("instr") or "").strip() == iid and to_int(rec.get("day"), 0) == d]
            if s in unavailable:
                issues.append(f"{code}: Eğitmen {iid} için G{d}/S{s} uygun değil.")

    # Conflict / group violations
    conflict_pairs = set()
    for a, b, _w in appstate.conflicts:
        a2, b2 = norm_code(a), norm_code(b)
        if a2 and b2:
            conflict_pairs.add((a2, b2))
            conflict_pairs.add((b2, a2))

    diffslot_pairs = set()
    for g in appstate.group_constraints:
        if str(g.get("type") or "") == "DifferentSlot":
            L = [norm_code(x) for x in (g.get("courses") or []) if norm_code(x)]
            L = [x for x in L if x in appstate.courses]
            for i in range(len(L)):
                for j in range(i + 1, len(L)):
                    diffslot_pairs.add((L[i], L[j]))
                    diffslot_pairs.add((L[j], L[i]))

    for a, b in conflict_pairs | diffslot_pairs:
        if a in placed and b in placed and placed[a] == placed[b]:
            reason_parts = []
            if (a, b) in conflict_pairs and (b, a) in conflict_pairs:
                reason_parts.append("dersler arasında doğrudan çakışma kaydı var")
            if (a, b) in diffslot_pairs:
                reason_parts.append("DifferentSlot grubu tarafından farklı slotta olmaları bekleniyor")
            reason = "; ".join(reason_parts) if reason_parts else "çakışma veya grup kısıtı nedeniyle"
            issues.append(
                f"{a} ve {b}: aynı G{placed[a][0]}/S{placed[a][1]} slotunda yerleşti. "
                f"Bu, {reason}."
            )

    # Slot capacity issues
    for (d, s), load in slot_load.items():
        cap = appstate.slot_caps.get((d, s), {"min": 0, "max": MAX_CAP})
        capmin = to_int(cap.get("min"), 0)
        capmax = to_int(cap.get("max"), MAX_CAP)
        if capmin > capmax:
            capmin, capmax = capmax, capmin
        if load < capmin:
            issues.append(f"G{d}/S{s}: Yük {load} < min kapasite {capmin}.")
        if capmax < MAX_CAP and load > capmax:
            issues.append(f"G{d}/S{s}: Yük {load} > max kapasite {capmax}.")

    return issues


# =========================
# CP-SAT solve relaxed (grup/çakışma constraints olmadan)
# =========================

def cpsat_solve_relaxed(courses: List[str], n_days: int, spd: int, allowed: Dict[str, List[int]], 
                       seed_map: Dict[str, Tuple[int, int, bool]]) -> Tuple[bool, str]:
    """Conflict ve group constraints olmadan basit atama yap"""
    # NOT: reset_results() çağrıyıcı tarafında zaten yapılmış
    P = n_days * spd
    
    model = cp_model.CpModel()
    
    t: Dict[str, cp_model.IntVar] = {}
    un: Dict[str, cp_model.BoolVar] = {}
    x: Dict[Tuple[str, int], cp_model.BoolVar] = {}
    
    for c in courses:
        t[c] = model.NewIntVar(0, P, f"t_{c}")
        un[c] = model.NewBoolVar(f"un_{c}")
        
        model.Add(t[c] == 0).OnlyEnforceIf(un[c])
        model.Add(t[c] != 0).OnlyEnforceIf(un[c].Not())
        
        domain_bools = []
        b0 = model.NewBoolVar(f"x_{c}_0")
        x[(c, 0)] = b0
        model.Add(t[c] == 0).OnlyEnforceIf(b0)
        model.Add(t[c] != 0).OnlyEnforceIf(b0.Not())
        domain_bools.append(b0)
        
        for p in range(1, P + 1):
            bp = model.NewBoolVar(f"x_{c}_{p}")
            x[(c, p)] = bp
            model.Add(t[c] == p).OnlyEnforceIf(bp)
            model.Add(t[c] != p).OnlyEnforceIf(bp.Not())
            domain_bools.append(bp)
        
        model.Add(sum(domain_bools) == 1)
    
    # Uygun slot constraints (allowed)
    for c in courses:
        allowed_set = set(allowed.get(c, []))
        for p in range(1, P + 1):
            if p not in allowed_set:
                model.Add(t[c] != p).OnlyEnforceIf(un[c].Not())
    
    # Capacity kapalı (skip)
    
    # Active expr - simplified
    active_expr = {}
    for c in courses:
        ms = to_int(appstate.courses[c].get("multi_slots"), 1)
        for p in range(1, P + 1):
            terms_c = []
            for off in range(ms):
                p_start = p - off
                if p_start >= 1 and (c, p_start) in x:
                    terms_c.append(x[(c, p_start)])
            active_expr[(c, p)] = LinearExpr.Sum(terms_c) if terms_c else LinearExpr.Constant(0)
    
    # Seed constraints
    for c, (d, sl, lock) in seed_map.items():
        if c not in t:
            continue
        fp = period_index(d, sl, spd)
        if lock:
            model.Add(t[c] == fp).OnlyEnforceIf(un[c].Not())
    
    # Objective: penalize unassigned önemlice
    PEN_UN = 10**12
    terms = [PEN_UN * x[(c, 0)] for c in courses]
    
    model.Minimize(LinearExpr.Sum(terms))
    
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 10.0
    solver.parameters.num_search_workers = 4
    
    status = solver.Solve(model)
    
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        msg = f"Çözüm bulunamadı. {len([c for c in courses if allowed.get(c)])} ders uygun slot var."
        return False, msg
    
    assigned_map: Dict[str, Tuple[int, int]] = {}
    unassigned: List[str] = []
    for c in courses:
        p = to_int(solver.Value(t[c]), 0)
        if p <= 0:
            unassigned.append(c)
        else:
            assigned_map[c] = period_to_day_slot(p, spd)
    
    expanded: Dict[str, Tuple[int, int]] = {}
    for c, (d, s) in assigned_map.items():
        expanded[c] = (d, s)
        ms = to_int(appstate.courses[c].get("multi_slots"), 1)
        if ms > 1:
            for off in range(1, ms):
                expanded[f"{c}__extra__{off}"] = (d, s + off)
    
    appstate.final = expanded
    
    appstate.final_unassigned = []
    for c in unassigned:
        cd = appstate.courses.get(c, {})
        alts: List[Tuple[int, int]] = []
        allowed_for_c = allowed.get(c, [])
        pref_days = [to_int(x, 0) for x in (cd.get("preferred_days") or []) if 1 <= to_int(x, 0) <= n_days]
        if not allowed_for_c:
            if pref_days:
                reason = f"Tercih günlerinde uygun slot yok ({pref_days})"
            else:
                reason = "Uygun slot yok"
        else:
            reason = "Kapasite/Çakışma: görmek için alternatif slotlara bak"
            for p in allowed_for_c[:3]:
                d, s = period_to_day_slot(p, spd)
                alts.append((d, s))

        appstate.final_unassigned.append((c, reason, alts))
    
    msg = f"Çözüm tamamlandı. {len(assigned_map)}/{len(courses)} ders atandı."
    msg += " Tercih günleri hard constraint olarak uygulandı. Tercih dışına atama yapılmadı."
    return True, msg


# =========================
# CP-SAT solve (UNASSIGNED + soft conflicts)
# =========================

def cpsat_solve(time_limit_sec: int = 20) -> Tuple[bool, str]:
    reset_results()
    appstate.last_diagnostics.append("Tercih günleri hard constraint olarak uygulanıyor.")
    if not HAS_ORTOOLS:
        return False, "OR-Tools yüklü değil. pip install ortools"

    n_days = to_int(appstate.calendar.get("n_days"), 9)
    spd = to_int(appstate.calendar.get("slots_per_day"), 4)
    P = n_days * spd

    courses = sorted(appstate.courses.keys())
    if not courses:
        return True, "Ders yok."

    model = cp_model.CpModel()

    t: Dict[str, cp_model.IntVar] = {}
    un: Dict[str, cp_model.BoolVar] = {}
    x: Dict[Tuple[str, int], cp_model.BoolVar] = {}

    for c in courses:
        t[c] = model.NewIntVar(0, P, f"t_{c}")
        un[c] = model.NewBoolVar(f"un_{c}")

        model.Add(t[c] == 0).OnlyEnforceIf(un[c])
        model.Add(t[c] != 0).OnlyEnforceIf(un[c].Not())

        domain_bools = []
        b0 = model.NewBoolVar(f"x_{c}_0")
        x[(c, 0)] = b0
        model.Add(t[c] == 0).OnlyEnforceIf(b0)
        model.Add(t[c] != 0).OnlyEnforceIf(b0.Not())
        domain_bools.append(b0)

        for p in range(1, P + 1):
            bp = model.NewBoolVar(f"x_{c}_{p}")
            x[(c, p)] = bp
            model.Add(t[c] == p).OnlyEnforceIf(bp)
            model.Add(t[c] != p).OnlyEnforceIf(bp.Not())
            domain_bools.append(bp)

        model.Add(sum(domain_bools) == 1)

    # Instructor unavailability lookup
    unav: Dict[Tuple[str, int], set[int]] = defaultdict(set)
    for rec in appstate.instructor_unavailability:
        iid = str(rec.get("instr") or "").strip()
        d = to_int(rec.get("day"), 0)
        s = to_int(rec.get("slot"), 0)
        if iid and 1 <= d <= n_days and 1 <= s <= spd:
            unav[(iid, d)].add(s)

    allowed: Dict[str, List[int]] = {}
    for c in courses:
        cd = appstate.courses[c]
        ms = to_int(cd.get("multi_slots"), 1)
        forb = set(to_int(x, 0) for x in (cd.get("forbidden_days") or []))
        pref_days = [to_int(x, 0) for x in (cd.get("preferred_days") or []) if 1 <= to_int(x, 0) <= n_days]
        if pref_days:
            pref_days = sorted(set(pref_days))
        instrs = [str(i).strip() for i in (cd.get("instructors") or []) if str(i).strip()]
        fd = to_int(cd.get("fixed_day"), 0)
        fs = to_int(cd.get("fixed_slot"), 0)
        fixed_fp = None
        if fd > 0 and fs > 0 and 1 <= fd <= n_days and 1 <= fs <= spd:
            fixed_fp = period_index(fd, fs, spd)

        a = []
        for d in range(1, n_days + 1):
            # Sabit gün varsa sadece o günü dene; yoksa tercih günlerine bak
            if fd > 0 and d != fd:
                continue
            if fd == 0 and pref_days and d not in pref_days:
                continue
            if d in forb:
                continue

            for s in range(1, spd + 1):
                # Sabit slot varsa sadece o slot
                if fs > 0 and s != fs:
                    continue
                # Fiziksel sığma: ders gün sonunu aşmamalı
                if s + ms - 1 > spd:
                    continue

                ok = True
                # Eğitmen uygunluğu: dersin süreceği tüm slotlarda kontrol et
                for off in range(ms):
                    curr_s = s + off
                    for iid in instrs:
                        if curr_s in unav.get((iid, d), set()):
                            ok = False
                            break
                    if not ok:
                        break
                if not ok:
                    continue
                a.append(period_index(d, s, spd))

        # Eğer sabit gün+slot varsa ve allowed listesine giremediyse zorla ekle
        # (örn. eğitmen müsait değilse giremeyebilir; model çözümsüz kalır ve diagnose uyarır)
        if fixed_fp is not None and fixed_fp not in a:
            a.insert(0, fixed_fp)
        allowed[c] = a
        allowed_set = set(a)
        for p in range(1, P + 1):
            if p not in allowed_set:
                model.Add(t[c] != p).OnlyEnforceIf(un[c].Not())

        # Sabit gün+slot olan dersler MUTLAKA atanmalı
        if fd > 0 and fs > 0:
            model.Add(un[c] == 0)
            model.Add(t[c] == fixed_fp)
        elif fixed_fp is not None:
            # Geriye dönük uyumluluk: sadece fixed_fp varsa ama hard zorunluluk değilse
            model.Add(t[c] == fixed_fp).OnlyEnforceIf(un[c].Not())

    # Hangi ders hangi periyotlarda AKTİF (yer kaplıyor) haritası
    active_expr = {}
    for c in courses:
        ms = to_int(appstate.courses[c].get("multi_slots"), 1)
        for p in range(1, P + 1):
            terms_c = []
            # Ders p periyodunda aktifse, (p - off) periyotlarından birinde başlamıştır.
            for off in range(ms):
                p_start = p - off
                if p_start >= 1 and (c, p_start) in x:
                    terms_c.append(x[(c, p_start)])
            active_expr[(c, p)] = LinearExpr.Sum(terms_c) if terms_c else LinearExpr.Constant(0)

    # 1. Kapasite (Hard Constraint) Güncellemesi
    size_map = {c: to_int(appstate.courses[c].get("size"), 0) for c in courses}
    
    # Seeds: build seed_map for later use
    seed_map: Dict[str, Tuple[int, int, bool]] = {}
    for s in appstate.seeds:
        c = norm_code(s.get("course"))
        d = to_int(s.get("day"), 0)
        sl = to_int(s.get("slot"), 0)
        lock = to_bool(s.get("lock"))
        if c and 1 <= d <= n_days and 1 <= sl <= spd:
            seed_map[c] = (d, sl, lock)
    for d in range(1, n_days + 1):
        for s in range(1, spd + 1):
            p = period_index(d, s, spd)
            cap = appstate.slot_caps.get((d, s), {"min": 0, "max": MAX_CAP})
            capmax = to_int(cap.get("max"), MAX_CAP)
            if capmax < MAX_CAP:
                # Sadece başlayanlar değil, o an AKTİF olan tüm derslerin kapasitesi toplanır
                load_expr = LinearExpr.Sum([size_map[c] * active_expr[(c, p)] for c in courses])
                model.Add(load_expr <= capmax)

    # Group constraints hard when both assigned
    def both_assigned(a: str, b: str) -> cp_model.BoolVar:
        bb = model.NewBoolVar(f"both_{a}_{b}")
        model.AddBoolAnd([un[a].Not(), un[b].Not()]).OnlyEnforceIf(bb)
        model.AddBoolOr([un[a], un[b]]).OnlyEnforceIf(bb.Not())
        return bb

    for g in appstate.group_constraints:
        gtype = str(g.get("type") or "")
        members = [norm_code(x) for x in (g.get("courses") or [])]
        members = [m for m in members if m in appstate.courses]
        if len(members) < 2:
            continue
        for i in range(len(members)):
            for j in range(i + 1, len(members)):
                a, b = members[i], members[j]
                if a not in t or b not in t:
                    continue
                ba = both_assigned(a, b)
                if gtype == "SameSlot":
                    model.Add(t[a] == t[b]).OnlyEnforceIf(ba)
                elif gtype == "DifferentSlot":
                    model.Add(t[a] != t[b]).OnlyEnforceIf(ba)
                elif gtype == "SameDay":
                    da = model.NewIntVar(1, n_days, f"day_{a}_{i}_{j}")
                    db = model.NewIntVar(1, n_days, f"day_{b}_{i}_{j}")
                    tuples = [(p, period_to_day_slot(p, spd)[0]) for p in range(1, P + 1)]
                    model.AddAllowedAssignments([t[a], da], tuples).OnlyEnforceIf(un[a].Not())
                    model.AddAllowedAssignments([t[b], db], tuples).OnlyEnforceIf(un[b].Not())
                    model.Add(da == db).OnlyEnforceIf(ba)
                elif gtype == "DifferentDay":
                    da = model.NewIntVar(1, n_days, f"day_{a}_D_{i}_{j}")
                    db = model.NewIntVar(1, n_days, f"day_{b}_D_{i}_{j}")
                    tuples = [(p, period_to_day_slot(p, spd)[0]) for p in range(1, P + 1)]
                    model.AddAllowedAssignments([t[a], da], tuples).OnlyEnforceIf(un[a].Not())
                    model.AddAllowedAssignments([t[b], db], tuples).OnlyEnforceIf(un[b].Not())
                    model.Add(da != db).OnlyEnforceIf(ba)

    # Seeds: hard/soft constraints applied below
    for c, (d, sl, lock) in seed_map.items():
        if c not in t:
            continue
        fp = period_index(d, sl, spd)
        if lock:
            model.Add(t[c] == fp).OnlyEnforceIf(un[c].Not())

    # Objective
    PEN_UN = 10**12
    terms = []
    for c in courses:
        terms.append(PEN_UN * x[(c, 0)])

    # Öğrenci çakışmaları (Soft Conflicts)
    seen = set()
    for a, b, w in appstate.conflicts:
        a2, b2 = norm_code(a), norm_code(b)
        if not a2 or not b2 or a2 == b2 or a2 not in t or b2 not in t:
            continue
        key = (a2, b2) if a2 < b2 else (b2, a2)
        if key in seen:
            continue
        seen.add(key)
        w2 = max(1, to_int(w, 1))
        
        overlap_bools = []
        for p in range(1, P + 1):
            # İki ders de o periyotta aktif mi?
            ov_p = model.NewBoolVar(f"ov_{a2}_{b2}_{p}")
            # Kesişme formülü: ov_p = a_aktif AND b_aktif
            model.Add(ov_p <= active_expr[(a2, p)])
            model.Add(ov_p <= active_expr[(b2, p)])
            model.Add(ov_p >= active_expr[(a2, p)] + active_expr[(b2, p)] - 1)
            overlap_bools.append(ov_p)
                
        if not overlap_bools:
            continue
            
        viol = model.NewBoolVar(f"conf_{a2}_{b2}")
        model.AddBoolOr(overlap_bools).OnlyEnforceIf(viol)
        model.AddBoolAnd([ob.Not() for ob in overlap_bools]).OnlyEnforceIf(viol.Not())
        
        # Sadece iki ders de atanmışsa (unassigned değillerse) ceza yaz
        viol_assigned = model.NewBoolVar(f"conf_ass_{a2}_{b2}")
        model.AddBoolAnd([viol, un[a2].Not(), un[b2].Not()]).OnlyEnforceIf(viol_assigned)
        model.AddBoolOr([viol.Not(), un[a2], un[b2]]).OnlyEnforceIf(viol_assigned.Not())
        
        terms.append(w2 * 1000 * viol_assigned)

    for c in courses:
        cd = appstate.courses[c]
        fd = to_int(cd.get("fixed_day"), 0)
        fs = to_int(cd.get("fixed_slot"), 0)
        if fd == 0 and 1 <= fs <= spd:
            sc = model.NewIntVar(1, spd, f"slot_pref_{c}")
            tuples = [(p, period_to_day_slot(p, spd)[1]) for p in range(1, P + 1)]
            model.AddAllowedAssignments([t[c], sc], tuples).OnlyEnforceIf(un[c].Not())
            ok = model.NewBoolVar(f"slotok_{c}")
            model.Add(sc == fs).OnlyEnforceIf(ok)
            model.Add(sc != fs).OnlyEnforceIf(ok.Not())
            viol = model.NewBoolVar(f"slotviol_{c}")
            model.Add(viol <= un[c].Not())
            model.Add(viol <= ok.Not())
            model.Add(viol >= un[c].Not() + ok.Not() - 1)
            terms.append(50 * viol)

    # soft capacity min shortfall
    for d in range(1, n_days + 1):
        for s in range(1, spd + 1):
            p = period_index(d, s, spd)
            cap = appstate.slot_caps.get((d, s), {"min": 0, "max": MAX_CAP})
            capmin = to_int(cap.get("min"), 0)
            if capmin <= 0:
                continue
            load = model.NewIntVar(0, sum(size_map.values()) + 1, f"load_{p}")
            model.Add(load == sum(size_map[c] * x[(c, p)] for c in courses))
            short = model.NewIntVar(0, capmin, f"short_{p}")
            model.Add(short >= capmin - load)
            model.Add(short >= 0)
            terms.append(short)

    # soft unlocked seed deviation
    for c, (d, sl, lock) in seed_map.items():
        if c not in t or lock:
            continue
        fp = period_index(d, sl, spd)
        devi = model.NewBoolVar(f"seed_dev_{c}")
        model.Add(t[c] == fp).OnlyEnforceIf(devi.Not())
        model.Add(t[c] != fp).OnlyEnforceIf(devi)
        dv = model.NewBoolVar(f"seed_dev_ass_{c}")
        model.Add(dv <= devi)
        model.Add(dv <= un[c].Not())
        model.Add(dv >= devi + un[c].Not() - 1)
        terms.append(10_000 * dv)

    model.Minimize(LinearExpr.Sum(terms))

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = float(max(1, to_int(time_limit_sec, 20)))
    solver.parameters.num_search_workers = 8

    status = solver.Solve(model)
    
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        appstate.last_diagnostics.append("Çözüm bulunamadı. Tercih günleri ve diğer kısıtlar hard olarak uygulandı.")
        # Hangi derslerin hiç uygun slotu yok?
        zero_allowed = [c for c in courses if not allowed.get(c)]
        if zero_allowed:
            appstate.last_diagnostics.append(f"Hiç uygun slotu olmayan dersler ({len(zero_allowed)}): {', '.join(zero_allowed[:10])}")
        appstate.last_diagnostics.extend(diagnose_all_issues())
        n_diag = len(appstate.last_diagnostics)
        return False, f"Çözüm bulunamadı. Kısıtlar nedeniyle yerleştirme mümkün değil. ({n_diag} adet teşhis notu üretildi. Ekrandaki Uyarılar sekmesini kontrol edin.)"

    assigned_map: Dict[str, Tuple[int, int]] = {}
    unassigned: List[str] = []
    for c in courses:
        p = to_int(solver.Value(t[c]), 0)
        if p <= 0:
            unassigned.append(c)
        else:
            assigned_map[c] = period_to_day_slot(p, spd)

    # Expand multi_slots -> __extra__ entries for UI table
    expanded: Dict[str, Tuple[int, int]] = {}
    for c, (d, s) in assigned_map.items():
        expanded[c] = (d, s)
        ms = to_int(appstate.courses[c].get("multi_slots"), 1)
        if ms > 1:
            for off in range(1, ms):
                expanded[f"{c}__extra__{off}"] = (d, s + off)

    appstate.final = expanded

    appstate.final_unassigned = []
    for c in unassigned:
        cd = appstate.courses.get(c, {})
        alts: List[Tuple[int, int]] = []
        reason = "Bilinmeyen neden"
        
        # Analyize sebepleri
        allowed_for_c = allowed.get(c, [])
        fd = to_int(cd.get("fixed_day"), 0)
        fs = to_int(cd.get("fixed_slot"), 0)
        if fd > 0 and fs > 0 and 1 <= fd <= n_days and 1 <= fs <= spd:
            fixed_fp = period_index(fd, fs, spd)
            if fixed_fp not in allowed_for_c:
                allowed_for_c = [fixed_fp] + allowed_for_c
        if not allowed_for_c:
            # Hiç uygun slot yok
            forb = [to_int(x, 0) for x in cd.get("forbidden_days", [])]
            instrs = [str(i).strip() for i in (cd.get("instructors") or []) if str(i).strip()]
            if forb:
                reason = f"Engelli gün(ler): {forb}"
            elif instrs:
                reason = "Eğitmen uygunluğu sorunu"
            else:
                reason = "Uygun slot yok"
        else:
            # Uygun slotlar var ama seçilmedi
            reason = "Kapasite/Çakışma: görmek için alternatif slotlara bak"
            for p in allowed_for_c[:3]:
                d, s = period_to_day_slot(p, spd)
                alts.append((d, s))
        
        if not alts and allowed_for_c:
            for p in allowed_for_c[:3]:
                d, s = period_to_day_slot(p, spd)
                alts.append((d, s))
        
        appstate.final_unassigned.append((c, reason, alts))

    return True, f"Çözüm tamamlandı. {len(assigned_map)}/{len(courses)} ders yerleşti."


# =========================
# Export
# =========================

def build_excel() -> bytes:
    if not HAS_PANDAS:
        raise RuntimeError("pandas gerekli: pip install pandas")

    assign = appstate.final or appstate.preview
    if not assign:
        raise RuntimeError("Takvim yok.")

    cal = cal_obj()
    if not cal.start_date:
        raise RuntimeError("Excel için Sınav Başlangıç Tarihi gerekli.")
    start_date = datetime.strptime(str(cal.start_date), "%Y-%m-%d").date()

    slot_len = to_int(cal.slot_length_min, 60)
    day_start = hhmm_to_time(cal.day_start_time) or time(8, 30)

    grouped: Dict[str, List[Tuple[int, int]]] = defaultdict(list)
    for code, (d, s) in assign.items():
        grouped[base_code(code)].append((to_int(d, 0), to_int(s, 0)))

    rows = []
    for code, slots in sorted(grouped.items()):
        slots = sorted(slots)
        d0 = slots[0][0]
        exam_date = start_date + timedelta(days=d0 - 1)
        starts_ends = [day_slot_times(ss, day_start, slot_len) for _, ss in slots]
        st = min(starts_ends, key=lambda x: (x[0].hour, x[0].minute))[0]
        en = max(starts_ends, key=lambda x: (x[1].hour, x[1].minute))[1]
        cd = appstate.courses.get(code, {})
        rows.append({
            "Ders Kodu": code,
            "Ders Adı": str(cd.get("name") or ""),
            "Öğrenci": to_int(cd.get("size"), 0),
            "Gün": d0,
            "Tarih": exam_date.strftime("%Y-%m-%d"),
            "Saat": f"{time_to_hhmm(st)}-{time_to_hhmm(en)}",
            "Slotlar": ", ".join([f"G{d}/S{s}" for d, s in slots]),
        })

    df = pd.DataFrame(rows)  # type: ignore[name-defined]
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:  # type: ignore[name-defined]
        df.to_excel(writer, index=False, sheet_name="Schedule")
    bio.seek(0)
    return bio.read()


def build_ics() -> bytes:
    assign = appstate.final or appstate.preview
    if not assign:
        raise RuntimeError("Takvim yok.")
    cal = cal_obj()
    if not cal.start_date:
        raise RuntimeError("ICS için Sınav Başlangıç Tarihi gerekli.")
    start_date = datetime.strptime(str(cal.start_date), "%Y-%m-%d").date()

    slot_len = to_int(cal.slot_length_min, 60)
    day_start = hhmm_to_time(cal.day_start_time) or time(8, 30)

    grouped: Dict[str, List[Tuple[int, int]]] = defaultdict(list)
    for code, (d, s) in assign.items():
        if "__extra__" in code:
            continue
        grouped[base_code(code)].append((to_int(d, 0), to_int(s, 0)))

    lines = ["BEGIN:VCALENDAR", "VERSION:2.0", "PRODID:-//ExamScheduler//TR//EN"]
    uid_base = int(datetime.now().timestamp())

    def fmt(dt: datetime) -> str:
        return dt.strftime("%Y%m%dT%H%M%S")

    for i, (code, slots) in enumerate(sorted(grouped.items())):
        slots = sorted(slots)
        d0 = slots[0][0]
        exam_date = start_date + timedelta(days=d0 - 1)
        starts_ends = [day_slot_times(ss, day_start, slot_len) for _, ss in slots]
        st = min(starts_ends, key=lambda x: (x[0].hour, x[0].minute))[0]
        en = max(starts_ends, key=lambda x: (x[1].hour, x[1].minute))[1]

        dtstart = datetime.combine(exam_date, st)
        dtend = datetime.combine(exam_date, en)

        cd = appstate.courses.get(code, {})
        summary = f"{code} {cd.get('name','')}".strip()
        desc = "Slots: " + ", ".join([f"G{d}/S{s}" for d, s in slots])

        lines += [
            "BEGIN:VEVENT",
            f"UID:{uid_base+i}@exam-scheduler",
            f"DTSTAMP:{fmt(datetime.now())}",
            f"DTSTART:{fmt(dtstart)}",
            f"DTEND:{fmt(dtend)}",
            f"SUMMARY:{summary}",
            f"DESCRIPTION:{desc}",
            "END:VEVENT",
        ]

    lines.append("END:VCALENDAR")
    return ("\r\n".join(lines) + "\r\n").encode("utf-8")


# =========================
# Kontrol Analizi
# =========================

def run_kontrol_analizi() -> Dict[str, Any]:
    appstate.kontrol_results = {}
    appstate.degistir_onerileri = []

    assign = appstate.final or appstate.preview
    if not assign:
        return {}

    n_days = to_int(appstate.calendar.get("n_days"), 9)
    spd = to_int(appstate.calendar.get("slots_per_day"), 4)

    placed = {base_code(k): (to_int(v[0], 0), to_int(v[1], 0)) for k, v in assign.items() if "__extra__" not in k}

    conflict_pairs: set[Tuple[str, str]] = set()
    for a, b, _w in appstate.conflicts:
        a2, b2 = norm_code(a), norm_code(b)
        if a2 and b2:
            conflict_pairs.add((a2, b2))
            conflict_pairs.add((b2, a2))

    diffday_pairs: set[Tuple[str, str]] = set()
    diffslot_pairs: set[Tuple[str, str]] = set()
    for g in appstate.group_constraints:
        gtype = str(g.get("type") or "")
        members = [norm_code(x) for x in (g.get("courses") or []) if norm_code(x)]
        members = [m for m in members if m in appstate.courses]
        if len(members) < 2:
            continue
        for i in range(len(members)):
            for j in range(i + 1, len(members)):
                a, b = members[i], members[j]
                if gtype == "DifferentDay":
                    diffday_pairs.add((a, b))
                    diffday_pairs.add((b, a))
                elif gtype == "DifferentSlot":
                    diffslot_pairs.add((a, b))
                    diffslot_pairs.add((b, a))

    cakisma_ihlal: List[Dict[str, Any]] = []
    diffday_ihlal: List[Dict[str, Any]] = []
    diffslot_ihlal: List[Dict[str, Any]] = []

    seen_conflict: set[Tuple[str, str]] = set()
    for (a, b) in conflict_pairs:
        if a >= b:
            continue
        if a in placed and b in placed and placed[a] == placed[b]:
            key = (a, b)
            if key not in seen_conflict:
                seen_conflict.add(key)
                cakisma_ihlal.append({"ders1": a, "ders2": b, "gun": placed[a][0], "slot": placed[a][1]})

    seen_diffday: set[Tuple[str, str]] = set()
    for (a, b) in diffday_pairs:
        if a >= b:
            continue
        if a in placed and b in placed and placed[a][0] == placed[b][0]:
            key = (a, b)
            if key not in seen_diffday:
                seen_diffday.add(key)
                diffday_ihlal.append({"ders1": a, "ders2": b, "gun": placed[a][0], "slot1": placed[a][1], "slot2": placed[b][1]})

    seen_diffslot: set[Tuple[str, str]] = set()
    for (a, b) in diffslot_pairs:
        if a >= b:
            continue
        if a in placed and b in placed and placed[a] == placed[b]:
            key = (a, b)
            if key not in seen_diffslot:
                seen_diffslot.add(key)
                diffslot_ihlal.append({"ders1": a, "ders2": b, "gun": placed[a][0], "slot": placed[a][1]})

    eksik_dersler = [c for c in appstate.courses if c not in placed]

    degisecek_set: set[str] = set()
    for ih in cakisma_ihlal + diffday_ihlal + diffslot_ihlal:
        for code in (ih["ders1"], ih["ders2"]):
            degisecek_set.add(code)

    degisecek_dersler = sorted(degisecek_set)

    oneriler: List[Dict[str, Any]] = []

    # --- A) Yerlesen ama ihlal eden dersler icin oneri ---
    for code in degisecek_dersler:
        if code not in placed:
            continue
        cd, cs = placed[code]
        found = False
        for d in range(1, n_days + 1):
            for s in range(1, spd + 1):
                if d == cd and s == cs:
                    continue
                ok = True
                for other, (od, os) in placed.items():
                    if other == code:
                        continue
                    if (d, s) == (od, os):
                        if (code, other) in conflict_pairs or (other, code) in conflict_pairs:
                            ok = False
                            break
                        if (code, other) in diffslot_pairs or (other, code) in diffslot_pairs:
                            ok = False
                            break
                    if d == od:
                        if (code, other) in diffday_pairs or (other, code) in diffday_pairs:
                            ok = False
                            break
                if ok:
                    neden = ""
                    if any(ih.get("ders1") == code or ih.get("ders2") == code for ih in cakisma_ihlal):
                        neden = "Cakisma ihlali"
                    elif any(ih.get("ders1") == code or ih.get("ders2") == code for ih in diffday_ihlal):
                        neden = "DiffDay ihlali"
                    elif any(ih.get("ders1") == code or ih.get("ders2") == code for ih in diffslot_ihlal):
                        neden = "DiffSlot ihlali"
                    oneriler.append({
                        "code": code,
                        "old_slot": f"G{cd}/S{cs}",
                        "new_slot": f"G{d}/S{s}",
                        "new_day": d,
                        "new_slot_val": s,
                        "neden": neden,
                    })
                    found = True
                    break
            if found:
                break
        if not found:
            oneriler.append({
                "code": code,
                "old_slot": f"G{cd}/S{cs}",
                "new_slot": f"G{cd}/S{cs}",
                "new_day": cd,
                "new_slot_val": cs,
                "neden": "Uygun alternatif bulunamadi",
            })

    # --- B) Yerlesemeyen dersler icin de oneri uret ---
    for code in eksik_dersler:
        cd = appstate.courses.get(code, {})
        fd = to_int(cd.get("fixed_day"), 0)
        fs = to_int(cd.get("fixed_slot"), 0)
        forb = set(to_int(x, 0) for x in (cd.get("forbidden_days") or []))
        pref_days = [to_int(x, 0) for x in (cd.get("preferred_days") or []) if 1 <= to_int(x, 0) <= n_days]

        found = False
        days_to_try = pref_days + [d for d in range(1, n_days + 1) if d not in pref_days]
        for d in days_to_try:
            if d in forb:
                continue
            for s in range(1, spd + 1):
                if fd > 0 and d != fd:
                    continue
                if fs > 0 and s != fs:
                    continue
                ok = True
                for other, (od, os) in placed.items():
                    if (d, s) == (od, os):
                        if (code, other) in conflict_pairs or (other, code) in conflict_pairs:
                            ok = False
                            break
                        if (code, other) in diffslot_pairs or (other, code) in diffslot_pairs:
                            ok = False
                            break
                    if d == od:
                        if (code, other) in diffday_pairs or (other, code) in diffday_pairs:
                            ok = False
                            break
                if ok:
                    oneriler.append({
                        "code": code,
                        "old_slot": "YERLESEMEDI",
                        "new_slot": f"G{d}/S{s}",
                        "new_day": d,
                        "new_slot_val": s,
                        "neden": "Yerlesemeyen ders - onerilen slot",
                    })
                    found = True
                    break
            if found:
                break
        if not found:
            oneriler.append({
                "code": code,
                "old_slot": "YERLESEMEDI",
                "new_slot": "YOK",
                "new_day": 0,
                "new_slot_val": 0,
                "neden": "Yerlesemeyen ders - hic uygun slot yok",
            })

    appstate.kontrol_results = {
        "cakisma_ihlal": cakisma_ihlal,
        "diffday_ihlal": diffday_ihlal,
        "diffslot_ihlal": diffslot_ihlal,
        "eksik_dersler": eksik_dersler,
        "degisecek_dersler": degisecek_dersler,
    }
    appstate.degistir_onerileri = oneriler
    return appstate.kontrol_results


# =========================
# Routes - MUST match HTML
# =========================

@app.route("/")
def index():
    cal = cal_obj()
    active_tab = request.args.get("tab", "calendar")
    edit_course_code = request.args.get("edit_course")
    edit_conflict_idx = request.args.get("edit_conflict")
    if edit_conflict_idx is not None:
        try:
            edit_conflict_idx = int(edit_conflict_idx)
        except Exception:
            edit_conflict_idx = None

    assign = appstate.final or appstate.preview

    slot_load: Dict[Tuple[int, int], int] = defaultdict(int)
    explicit_extra_bases = {base_code(code) for code in assign.keys() if "__extra__" in code}
    for code, (d, s) in assign.items():
        d_i = to_int(d, 0)
        s_i = to_int(s, 0)
        if d_i <= 0 or s_i <= 0:
            continue
        base = base_code(code)
        size = to_int(appstate.courses.get(base, {}).get("size"), 0)
        slot_load[(d_i, s_i)] += size
        if "__extra__" not in code:
            ms = to_int(appstate.courses.get(base, {}).get("multi_slots"), 1)
            if ms > 1 and base not in explicit_extra_bases:
                for off in range(1, ms):
                    slot_load[(d_i, s_i + off)] += size

    def heat_cell(d: int, s: int) -> Tuple[str, int, int]:
        d = to_int(d, 0)
        s = to_int(s, 0)
        cap = appstate.slot_caps.get((d, s), {"min": 0, "max": MAX_CAP})
        capmax = to_int(cap.get("max"), MAX_CAP)
        load = to_int(slot_load.get((d, s), 0), 0)
        denom = capmax if capmax < MAX_CAP else max(1, max(slot_load.values(), default=1))
        color = make_heat(load, denom)
        return color, load, capmax  # HTML expects this order

    def day_time(d: int, s: int) -> Tuple[str, str]:
        st0 = hhmm_to_time(cal.day_start_time) or time(8, 30)
        st, en = day_slot_times(to_int(s, 1), st0, to_int(cal.slot_length_min, 60))
        return time_to_hhmm(st), time_to_hhmm(en)

    check_fixed_ok = True
    for c, cd in appstate.courses.items():
        fd = to_int(cd.get("fixed_day"), 0)
        fs = to_int(cd.get("fixed_slot"), 0)
        if fd > 0 and fs > 0 and c in assign:
            if (to_int(assign[c][0], 0), to_int(assign[c][1], 0)) != (fd, fs):
                check_fixed_ok = False
                break

    check_caps_ok = True
    for (d, s), load in slot_load.items():
        cap = appstate.slot_caps.get((d, s), {"min": 0, "max": MAX_CAP})
        capmin = to_int(cap.get("min"), 0)
        capmax = to_int(cap.get("max"), MAX_CAP)
        if capmin > capmax:
            capmin, capmax = capmax, capmin
        if load < capmin or load > capmax:
            check_caps_ok = False
            break

    check_groups_ok = True
    placed = {base_code(k): (to_int(v[0], 0), to_int(v[1], 0)) for k, v in assign.items() if "__extra__" not in k}

    diffslot_pairs = set()
    for g in appstate.group_constraints:
        if str(g.get("type") or "") == "DifferentSlot":
            L = [norm_code(x) for x in (g.get("courses") or []) if norm_code(x)]
            L = [x for x in L if x in appstate.courses]
            for i in range(len(L)):
                for j in range(i + 1, len(L)):
                    diffslot_pairs.add((L[i], L[j]))
                    diffslot_pairs.add((L[j], L[i]))

    conf_pairs = set()
    for a, b, _w in appstate.conflicts:
        a2, b2 = norm_code(a), norm_code(b)
        if a2 and b2:
            conf_pairs.add((a2, b2))
            conf_pairs.add((b2, a2))
    conf_pairs |= diffslot_pairs

    for (a, b) in conf_pairs:
        if a in placed and b in placed and placed[a] == placed[b]:
            check_groups_ok = False
            break

    courses_wrapped: Dict[str, SimpleNamespace] = {code: course_obj(code, cd) for code, cd in appstate.courses.items()}
    edit_data = courses_wrapped.get(edit_course_code) if edit_course_code else None

    diagnostics = diagnose_all_issues() + list(appstate.last_diagnostics)
    assignment_issues = analyze_assignment_issues(assign, slot_load)
    solve_notes = list(appstate.last_diagnostics)

    # Atanamayan derslerin çakışma yaptığı dersleri bul
    conflict_pairs: Dict[str, set] = defaultdict(set)
    for a, b, _w in appstate.conflicts:
        a2, b2 = norm_code(a), norm_code(b)
        if a2 and b2:
            conflict_pairs[a2].add(b2)
            conflict_pairs[b2].add(a2)

    preview_unassigned_conflicts: Dict[str, List[str]] = {}
    for code, reason, alts in appstate.preview_unassigned:
        conflicting = sorted(conflict_pairs.get(norm_code(code), set()))
        preview_unassigned_conflicts[code] = conflicting

    final_unassigned_conflicts: Dict[str, List[str]] = {}
    for code, reason, alts in appstate.final_unassigned:
        conflicting = sorted(conflict_pairs.get(norm_code(code), set()))
        final_unassigned_conflicts[code] = conflicting

    return render_template(
        "index.html",
        cal=cal,
        courses=courses_wrapped,
        conflicts=appstate.conflicts,
        group_constraints=group_constraints_obj(),
        instructors=instructors_obj(),
        instructor_unavailability=unavailability_obj(),
        slot_caps=slot_caps_obj(),
        seeds=seeds_obj(),
        preview=appstate.preview,
        preview_unassigned=appstate.preview_unassigned,
        preview_unassigned_conflicts=preview_unassigned_conflicts,
        final=appstate.final,
        final_unassigned=appstate.final_unassigned,
        final_unassigned_conflicts=final_unassigned_conflicts,
        has_ortools=HAS_ORTOOLS,
        heat_cell=heat_cell,
        day_time=day_time,
        check_fixed_ok=check_fixed_ok,
        check_caps_ok=check_caps_ok,
        check_groups_ok=check_groups_ok,
        active_tab=active_tab,
        edit_course_code=edit_course_code,
        edit_data=edit_data,
        edit_conflict_idx=edit_conflict_idx,
        diagnostics=diagnostics,
        assignment_issues=assignment_issues,
        solve_notes=solve_notes,
        kontrol_results=appstate.kontrol_results,
        degistir_onerileri=appstate.degistir_onerileri,
    )


@app.route("/save_calendar", methods=["POST"])
def save_calendar():
    c = appstate.calendar
    c["n_days"] = clamp_int(request.form.get("n_days"), 1, 365) or to_int(c.get("n_days"), 9)
    c["slots_per_day"] = clamp_int(request.form.get("slots_per_day"), 1, 24) or to_int(c.get("slots_per_day"), 4)
    c["slot_length_min"] = clamp_int(request.form.get("slot_length_min"), 15, 600) or to_int(c.get("slot_length_min"), 60)
    c["buffer_minutes"] = clamp_int(request.form.get("buffer_minutes"), 0, 600) or 0
    c["day_start_time"] = (request.form.get("day_start_time") or c.get("day_start_time") or "08:30").strip()
    c["day_end_time"] = (request.form.get("day_end_time") or c.get("day_end_time") or "18:00").strip()
    c["last_slot_forbidden_for_three_hour"] = ("last_slot_forbidden_for_three_hour" in request.form)
    c["start_date"] = request.form.get("start_date") or None
    c["end_date"] = request.form.get("end_date") or None

    n = to_int(c["n_days"], 9)
    spd = to_int(c["slots_per_day"], 4)
    for d in range(1, n + 1):
        for s in range(1, spd + 1):
            cap = appstate.slot_caps.get((d, s), {"min": 0, "max": MAX_CAP})
            appstate.slot_caps[(d, s)] = {"min": to_int(cap.get("min"), 0), "max": to_int(cap.get("max"), MAX_CAP)}

    reset_results()
    flash("Takvim ayarları kaydedildi.", "success")
    return same_tab(request.args.get("tab", "calendar"))


@app.route("/reset_all")
def reset_all():
    reset_all_state()
    flash("Tüm veriler sıfırlandı.", "warning")
    return same_tab(request.args.get("tab", "calendar"))


@app.route("/upload_courses_excel", methods=["POST"])
def upload_courses_excel():
    f = request.files.get("file")
    if not f:
        flash("Dosya seçiniz.", "danger")
        return same_tab(request.args.get("tab", "courses"))
    try:
        load_courses_excel(f.read())
    except Exception as e:
        flash(f"Excel okunamadı: {e}", "danger")
    return same_tab(request.args.get("tab", "courses"))


@app.route("/add_course", methods=["POST"])
def add_course():
    code = norm_code(request.form.get("code"))
    if not code:
        flash("Ders kodu zorunlu.", "danger")
        return same_tab(request.args.get("tab", "courses"))

    n_days = to_int(appstate.calendar.get("n_days"), 9)
    spd = to_int(appstate.calendar.get("slots_per_day"), 4)
    slot_len = to_int(appstate.calendar.get("slot_length_min"), 60)

    name = str(request.form.get("name") or "").strip()
    duration = to_int(request.form.get("duration"), slot_len)
    size = to_int(request.form.get("size"), 0)
    multi_slots = to_int(request.form.get("multi_slots"), max(1, int(math.ceil(duration / max(1, slot_len)))))

    three_hour = ("three_hour" in request.form)
    if three_hour:
        duration = 180
        if not request.form.get("multi_slots"):
            multi_slots = max(1, int(math.ceil(duration / max(1, slot_len))))

    preferred = parse_days(request.form.get("preferred_days"), n_days)
    forbidden = parse_days(request.form.get("forbidden_days"), n_days)

    fixed_day = clamp_int(request.form.get("fixed_day"), 1, n_days)
    fixed_slot = clamp_int(request.form.get("fixed_slot"), 1, spd)

    instructors = parse_ids(request.form.get("instructor_ids"))

    appstate.courses[code] = {
        "code": code,
        "name": name,
        "duration": duration,
        "size": size,
        "preferred_days": preferred,
        "forbidden_days": forbidden,
        "fixed_day": fixed_day,
        "fixed_slot": fixed_slot,
        "multi_slots": multi_slots,
        "three_hour": three_hour,
        "instructors": instructors,
    }
    reset_results()
    flash("Ders eklendi.", "success")
    return same_tab(request.args.get("tab", "courses"))


@app.route("/update_course", methods=["POST"])
def update_course():
    code = norm_code(request.form.get("code"))
    if not code or code not in appstate.courses:
        flash("Ders bulunamadı.", "danger")
        return same_tab(request.args.get("tab", "courses"))

    n_days = to_int(appstate.calendar.get("n_days"), 9)
    spd = to_int(appstate.calendar.get("slots_per_day"), 4)
    slot_len = to_int(appstate.calendar.get("slot_length_min"), 60)

    cd = appstate.courses[code]
    cd["name"] = str(request.form.get("name") or "").strip()
    cd["duration"] = to_int(request.form.get("duration"), slot_len)
    cd["size"] = to_int(request.form.get("size"), 0)
    cd["multi_slots"] = to_int(request.form.get("multi_slots"), max(1, int(math.ceil(to_int(cd["duration"], slot_len) / max(1, slot_len)))))

    three_hour = ("three_hour" in request.form)
    cd["three_hour"] = three_hour
    if three_hour:
        cd["duration"] = 180
        if not request.form.get("multi_slots"):
            cd["multi_slots"] = max(1, int(math.ceil(180 / max(1, slot_len))))

    cd["preferred_days"] = parse_days(request.form.get("preferred_days"), n_days)
    cd["forbidden_days"] = parse_days(request.form.get("forbidden_days"), n_days)
    cd["fixed_day"] = clamp_int(request.form.get("fixed_day"), 1, n_days)
    cd["fixed_slot"] = clamp_int(request.form.get("fixed_slot"), 1, spd)
    cd["instructors"] = parse_ids(request.form.get("instructor_ids"))

    reset_results()
    flash("Ders güncellendi.", "success")
    return same_tab(request.args.get("tab", "courses"))


@app.route("/del_course")
def del_course():
    code = norm_code(request.args.get("code"))
    if code in appstate.courses:
        del appstate.courses[code]
        reset_results()
        flash("Ders silindi.", "warning")
    return same_tab(request.args.get("tab", "courses"))


@app.route("/upload_conflicts_csv", methods=["POST"])
def upload_conflicts_csv():
    f = request.files.get("csv_file") or request.files.get("file")
    if not f:
        flash("CSV seçiniz.", "danger")
        return same_tab(request.args.get("tab", "conflicts"))
    try:
        load_conflicts_csv(f.read())
    except Exception as e:
        flash(f"CSV okunamadı: {e}", "danger")
    return same_tab(request.args.get("tab", "conflicts"))


@app.route("/add_conflict", methods=["POST"])
def add_conflict():
    a = norm_code(request.form.get("c1"))
    b = norm_code(request.form.get("c2"))
    w = to_int(request.form.get("w"), 1)
    if not a or not b or a == b:
        flash("Çakışma için iki farklı ders giriniz.", "danger")
        return same_tab(request.args.get("tab", "conflicts"))
    # Only add conflicts when both courses exist in the system
    if a not in appstate.courses or b not in appstate.courses:
        flash("Çakışma eklenmedi: Bir veya iki ders sistemde yok.", "warning")
        return same_tab(request.args.get("tab", "conflicts"))
    appstate.conflicts.append((a, b, max(1, w)))
    reset_results()
    flash("Çakışma eklendi.", "success")
    return same_tab(request.args.get("tab", "conflicts"))


@app.route("/update_conflict", methods=["POST"])
def update_conflict():
    idx = clamp_int(request.form.get("idx"), 0, MAX_CAP)
    if idx is None or idx < 0 or idx >= len(appstate.conflicts):
        flash("Çakışma bulunamadı.", "danger")
        return same_tab(request.args.get("tab", "conflicts"))
    a = norm_code(request.form.get("c1"))
    b = norm_code(request.form.get("c2"))
    w = to_int(request.form.get("w"), 1)
    if not a or not b or a == b:
        flash("Geçersiz çakışma.", "danger")
        return same_tab(request.args.get("tab", "conflicts"))
    appstate.conflicts[idx] = (a, b, max(1, w))
    reset_results()
    flash("Çakışma güncellendi.", "success")
    return same_tab(request.args.get("tab", "conflicts"))


@app.route("/del_conflict")
def del_conflict():
    a = request.args.get("a")
    b = request.args.get("b")
    w = request.args.get("w")
    if a is None or b is None:
        flash("Silme parametreleri eksik.", "danger")
        return same_tab(request.args.get("tab", "conflicts"))

    a2 = norm_code(a)
    b2 = norm_code(b)
    w2 = to_int(w, None) if w is not None else None  # type: ignore[arg-type]

    removed = False
    for i, (aa, bb, ww) in enumerate(list(appstate.conflicts)):
        if aa == a2 and bb == b2 and (w2 is None or ww == w2):
            appstate.conflicts.pop(i)
            removed = True
            break
    if removed:
        reset_results()
        flash("Çakışma silindi.", "warning")
    else:
        flash("Çakışma bulunamadı.", "warning")
    return same_tab(request.args.get("tab", "conflicts"))


@app.route("/upload_group_constraints_excel", methods=["POST"])
def upload_group_constraints_excel():
    if "xlsx_file" not in request.files:
        flash("Dosya seçilmedi.", "danger")
        return same_tab(request.args.get("tab", "conflicts"))
    file = request.files["xlsx_file"]
    if file.filename == "":
        flash("Dosya seçilmedi.", "danger")
        return same_tab(request.args.get("tab", "conflicts"))
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        flash("Lütfen Excel dosyası (.xlsx/.xls) yükleyiniz.", "danger")
        return same_tab(request.args.get("tab", "conflicts"))
    try:
        data = file.read()
        load_group_constraints_excel(data)
    except Exception as e:
        flash(f"Excel yükleme hatası: {str(e)}", "error")
    return same_tab(request.args.get("tab", "conflicts"))


@app.route("/add_group_constraint", methods=["POST"])
def add_group_constraint():
    gtype = str(request.form.get("type") or "").strip()
    courses_raw = request.form.get("courses") or ""
    members = [norm_code(x) for x in re.split(r"[,\s;]+", courses_raw) if norm_code(x)]
    members = [m for m in members if m in appstate.courses]
    if gtype not in ("SameDay", "DifferentDay", "SameSlot", "DifferentSlot") or len(members) < 2:
        flash("Grup türü ve en az 2 ders giriniz.", "danger")
        return same_tab(request.args.get("tab", "conflicts"))
    appstate.group_constraints.append({"type": gtype, "courses": members})
    reset_results()
    flash("Grup kısıtı eklendi.", "success")
    return same_tab(request.args.get("tab", "conflicts"))


@app.route("/del_group_constraint")
def del_group_constraint():
    idx = clamp_int(request.args.get("idx"), 0, MAX_CAP)
    if idx is not None and 0 <= idx < len(appstate.group_constraints):
        appstate.group_constraints.pop(idx)
        reset_results()
        flash("Grup kısıtı silindi.", "warning")
    return same_tab(request.args.get("tab", "conflicts"))


@app.route("/add_instructor", methods=["POST"])
def add_instructor():
    iid = str(request.form.get("instr_id") or "").strip()
    name = str(request.form.get("name") or "").strip()
    email = str(request.form.get("email") or "").strip()
    if not iid:
        flash("Eğitmen ID gerekli.", "danger")
        return same_tab(request.args.get("tab", "instructors"))
    appstate.instructors[iid] = {"name": name, "email": email}
    flash("Eğitmen eklendi.", "success")
    return same_tab(request.args.get("tab", "instructors"))


@app.route("/del_instructor")
def del_instructor():
    iid = str(request.args.get("instr_id") or "").strip()
    if iid in appstate.instructors:
        del appstate.instructors[iid]
        flash("Eğitmen silindi.", "warning")
    return same_tab(request.args.get("tab", "instructors"))


@app.route("/add_unavailability", methods=["POST"])
def add_unavailability():
    iid = str(request.form.get("instr") or "").strip()
    day = clamp_int(request.form.get("day"), 1, MAX_CAP)
    slot = clamp_int(request.form.get("slot"), 1, MAX_CAP)
    if not iid or not day or not slot:
        flash("Uygun olmama için ID/gün/slot giriniz.", "danger")
    else:
        appstate.instructor_unavailability.append({"instr": iid, "day": to_int(day, 0), "slot": to_int(slot, 0)})
        reset_results()
        flash("Uygun olmama eklendi.", "success")
    return same_tab(request.args.get("tab", "instructors"))


@app.route("/del_unavailability")
def del_unavailability():
    idx = clamp_int(request.args.get("idx"), 0, MAX_CAP)
    if idx is not None and 0 <= idx < len(appstate.instructor_unavailability):
        appstate.instructor_unavailability.pop(idx)
        reset_results()
        flash("Kayıt silindi.", "warning")
    return same_tab(request.args.get("tab", "instructors"))


@app.route("/bulk_set_caps", methods=["POST"])
def bulk_set_caps():
    max_all = request.form.get("max_all")
    min_all = request.form.get("min_all")
    n = to_int(appstate.calendar.get("n_days"), 9)
    spd = to_int(appstate.calendar.get("slots_per_day"), 4)

    mm = to_opt_int(min_all)
    mx = to_opt_int(max_all)

    for d in range(1, n + 1):
        for s in range(1, spd + 1):
            cap = appstate.slot_caps.get((d, s), {"min": 0, "max": MAX_CAP})
            if mm is not None:
                cap["min"] = to_int(mm, 0)
            if mx is not None:
                cap["max"] = to_int(mx, MAX_CAP)
            if to_int(cap["min"], 0) > to_int(cap["max"], MAX_CAP):
                cap["min"], cap["max"] = cap["max"], cap["min"]
            appstate.slot_caps[(d, s)] = {"min": to_int(cap["min"], 0), "max": to_int(cap["max"], MAX_CAP)}

    reset_results()
    flash("Toplu kapasite uygulandı.", "success")
    return same_tab(request.args.get("tab", "slotcaps"))


@app.route("/save_caps", methods=["POST"])
def save_caps():
    n = to_int(appstate.calendar.get("n_days"), 9)
    spd = to_int(appstate.calendar.get("slots_per_day"), 4)
    for d in range(1, n + 1):
        for s in range(1, spd + 1):
            mn = to_opt_int(request.form.get(f"min_{d}_{s}"))
            mx_raw = request.form.get(f"max_{d}_{s}")
            mx = to_opt_int(mx_raw) if mx_raw not in (None, "") else MAX_CAP
            mn_i = to_int(mn, 0)
            mx_i = to_int(mx, MAX_CAP)
            if mn_i > mx_i:
                mn_i, mx_i = mx_i, mn_i
            appstate.slot_caps[(d, s)] = {"min": mn_i, "max": mx_i}
    reset_results()
    flash("Kapasiteler kaydedildi.", "success")
    return same_tab(request.args.get("tab", "slotcaps"))


@app.route("/add_seed", methods=["POST"])
def add_seed():
    c = norm_code(request.form.get("course"))
    d = clamp_int(request.form.get("day"), 1, MAX_CAP)
    s = clamp_int(request.form.get("slot"), 1, MAX_CAP)
    lock = ("lock" in request.form)
    if not c or not d or not s:
        flash("Tohum için ders/gün/slot giriniz.", "danger")
    else:
        appstate.seeds.append({"course": c, "day": to_int(d, 0), "slot": to_int(s, 0), "lock": lock})
        reset_results()
        flash("Tohum eklendi.", "success")
    return same_tab(request.args.get("tab", "seeds"))


@app.route("/del_seed")
def del_seed():
    idx = clamp_int(request.args.get("idx"), 0, MAX_CAP)
    if idx is not None and 0 <= idx < len(appstate.seeds):
        appstate.seeds.pop(idx)
        reset_results()
        flash("Tohum silindi.", "warning")
    return same_tab(request.args.get("tab", "seeds"))


@app.route("/run_preview", methods=["POST"])
def run_preview():
    greedy_preview()
    flash("Önizleme çalıştırıldı.", "success")
    return same_tab(request.args.get("tab", "preview"))


@app.route("/run_solve", methods=["POST"])
def run_solve():
    tl = to_int(request.form.get("tl"), 20)
    ok, msg = cpsat_solve(time_limit_sec=tl)
    flash(msg, "success" if ok else "danger")
    # Cozum sonrasi otomatik kontrol
    if ok:
        run_kontrol_analizi()
    return same_tab(request.args.get("tab", "solve"))


@app.route("/run_kontrol", methods=["POST"])
def run_kontrol():
    run_kontrol_analizi()
    flash("Kontrol analizi tamamlandi.", "success")
    return same_tab(request.args.get("tab", "kontrol"))


@app.route("/apply_changes", methods=["POST"])
def apply_changes():
    changes = request.form.getlist("changes")
    if not changes:
        flash("Hicbir degisiklik secilmedi.", "warning")
        return same_tab(request.args.get("tab", "degistir"))

    updated = 0
    for ch in changes:
        parts = ch.split("|")
        if len(parts) != 3:
            continue
        code = norm_code(parts[0])
        try:
            new_day = int(parts[1])
            new_slot = int(parts[2])
        except Exception:
            continue
        if new_day <= 0 or new_slot <= 0:
            # Yerlesmeyen ders icin YOK onerisi secildiyse atla
            continue
        if code not in appstate.final and code not in appstate.preview:
            continue
        if code in appstate.final:
            appstate.final[code] = (new_day, new_slot)
            ms = to_int(appstate.courses.get(code, {}).get("multi_slots"), 1)
            if ms > 1:
                for off in range(1, ms):
                    extra_key = f"{code}__extra__{off}"
                    if extra_key in appstate.final:
                        appstate.final[extra_key] = (new_day, new_slot + off)
        if code in appstate.preview:
            appstate.preview[code] = (new_day, new_slot)
        updated += 1

    if updated:
        flash(f"{updated} ders guncellendi.", "success")
        # Degisiklik sonrasi otomatik kontrol
        run_kontrol_analizi()
    else:
        flash("Guncellenecek ders bulunamadi.", "warning")
    return same_tab(request.args.get("tab", "degistir"))


@app.route("/download_excel")
def download_excel():
    try:
        data = build_excel()
    except RuntimeError as e:
        flash(f"Excel indirilemedi: {e}", "danger")
        return same_tab(request.args.get("tab", "solve"))
    except Exception as e:
        flash(f"Excel indirilemedi: {e}", "danger")
        return same_tab(request.args.get("tab", "solve"))
    return send_file(io.BytesIO(data), as_attachment=True, download_name="schedule.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/download_ics")
def download_ics():
    try:
        data = build_ics()
    except Exception as e:
        flash(str(e), "danger")
        return same_tab(request.args.get("tab", "solve"))
    return send_file(io.BytesIO(data), as_attachment=True, download_name="schedule.ics", mimetype="text/calendar")


@app.route("/save_ics", methods=["POST"])
def save_ics_post():
    return download_ics()


reset_all_state()

if __name__ == "__main__":
    n = to_int(appstate.calendar.get("n_days"), 9)
    spd = to_int(appstate.calendar.get("slots_per_day"), 4)
    for d in range(1, n + 1):
        for s in range(1, spd + 1):
            _ = appstate.slot_caps[(d, s)]
    app.run(host="127.0.0.1", port=5000, debug=False)

